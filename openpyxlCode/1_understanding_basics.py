from openpyxl.workbook import Workbook
from openpyxl import load_workbook
from faker import Faker
from random import randint

def main():
    # wb = Workbook() # actual workbook
    # ws = wb.active # active sheet in actual workbook
    wb = load_workbook('dataExcel.xlsx') # for loading an existing file
    ws = wb.active
    # print(ws['A1'].value)
    # for printing a column
    # for cell in ws['C']:
    #     print(cell.value)
    # for specifying a range of columns
    range = ws['A1': 'P1']
    for cell in range:
        for value in cell:
            print(value.value)        
    # for targeting a specific cell
    cell_obj = ws.cell(
        row = 1,
        column = 1
    ).value
    print(cell_obj)
    # for number of rows and columns used in a sheet
    print(f'Rows: {ws.max_row} | Columns: {ws.max_column}')
    # printing all rows and columns
    # rows, cols = int(ws.max_row), int(ws.max_column)
    # print(f'rows: {rows} | cols: {cols}\nDatatype: {type(rows) | type(cols)}')
    # i = 1
    # while (i != 10):
    #     j = 1
    #     while (j != 15):
    #         print(
    #             ws.cell(
    #                 row = i,
    #                 column = j
    #             ).value, end=' | '
    #         )
    #         j += 1
    #     print('\n')
    #     i += 1
    # sliced_set = ws['A1': 'L5']
    # # print(sliced_set)
    # for cell1 in sliced_set:
    #     for cell2 in cell1:
    #         print(f'{cell2.value}', end=' | ')
    #     print()
    # writing data in a workbook
    # wbook = Workbook()
    # sheet_obj = wbook.active
    # sheet_obj['A1'].value = 'Value1'
    # sheet_obj['A2'].value = 'Value2'
    # wbook.save('savedFile.xlsx')
    # wbook = load_workbook('savedFile.xlsx')
    # sheet_obj = wbook.active
    # print(sheet_obj['A1'].value)
    # print(sheet_obj['A2'].value)
    # writing multiple rows into a workbook
    wbook = Workbook()
    sheet_obj = wbook.active
    headers = (
        'Name',
        'Address',
        'Email',
        'Phone Number',
        'Salary'
    )
    fake = Faker()
    data = (
        headers,
        (fake.name(), fake.address(), fake.email(), fake.phone_number(), randint(10000, 70000) ),
        (fake.name(), fake.address(), fake.email(), fake.phone_number(), randint(10000, 70000) ),
        (fake.name(), fake.address(), fake.email(), fake.phone_number(), randint(10000, 70000) ),
        (fake.name(), fake.address(), fake.email(), fake.phone_number(), randint(10000, 70000) ),
        (fake.name(), fake.address(), fake.email(), fake.phone_number(), randint(10000, 70000) ),
    )
    for cell in data:
        sheet_obj.append(cell)
    sheet_obj['A7'] = 'Total Salary:'
    sheet_obj['E7'] = '=SUM(E2:E6)'
    sheet_obj.column_dimensions['A'].width = 20
    sheet_obj.column_dimensions['B'].width = 50
    sheet_obj.column_dimensions['C'].width = 30
    sheet_obj.column_dimensions['D'].width = 20
    sheet_obj.column_dimensions['E'].width = 10
    sheet_obj.row_dimensions[1].height = 30
    sheet_obj.merge_cells('A8:E16')
    wbook.save('fileWithMultipleRows.xlsx')
    # reading the newly created file that has multiple rows
    wb = load_workbook('fileWithMultipleRows.xlsx')
    ws = wb.active
    i = 1
    rows, cols = ws.max_row, ws.max_column
    while (i != rows+1):
        j = 1
        while (j != cols+1):
            print(ws.cell(row=i, column=j).value, end=' | ')
            j += 1
        print()
        i += 1
if __name__ == '__main__':
    main()