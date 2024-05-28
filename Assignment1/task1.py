import openpyxl

def implement():
    numbers = []
    print('Enter 5 numbers')
    for _ in range(5):
        num = int(input('Enter a number: '))
        numbers.append(num)
    work_book_obj = openpyxl.Workbook() # load_workbook
    sheet_obj = work_book_obj.active
    for i in range(1, 6):
        cell = sheet_obj.cell(
            row = i,
            column = 1
        )        
        cell.value = numbers[i-1]
    cell = sheet_obj['A7'] # { a:{1: [], 2: [], 3: [], 4: [], 5: [], 6:[], 7: [] }
    cell.value = '=SUM(A1:A6)'
    avg_cell = sheet_obj['B7']
    avg_cell.value = '=AVERAGE(A1:A6)'
    sheet_obj.row_dimensions[7].height = 50
    sheet_obj.column_dimensions['B'].width = 50
    work_book_obj.save('task1File.xlsx')

def main():
    implement()

if __name__ == '__main__':
    main()