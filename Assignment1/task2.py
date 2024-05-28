import openpyxl

def merge_cells(file_path, start_cell, end_cell):
    try:
        wb = openpyxl.load_workbook(file_path)
        sheet = wb.active
        sheet.merge_cells(start_cell + ':' + end_cell)
        wb.save(file_path)
        print("Cells merged successfully!")
    except Exception as error:
        print("Error occurred:", str(error))

def main():
    file_path = input("Enter the path of the Excel file or press 1 for default file: ")
    if file_path == '1':
        file_path = 'task2File.xlsx'
    start_cell = input("Enter the starting cell (e.g., A1): ")
    end_cell = input("Enter the ending cell (e.g., B2): ")
    merge_cells(file_path, start_cell, end_cell)

if __name__ == "__main__":
    main()
